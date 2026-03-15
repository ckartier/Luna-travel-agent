#!/usr/bin/env python3
"""
Recalculate formulas in an Excel file using LibreOffice.
Usage: python recalc.py <excel_file> [timeout_seconds]
"""
import sys
import os
import json
import subprocess
import tempfile
import shutil
from pathlib import Path

def find_libreoffice():
    """Find LibreOffice binary."""
    paths = [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/libreoffice",
        "/usr/bin/soffice",
        shutil.which("libreoffice") or "",
        shutil.which("soffice") or "",
    ]
    for p in paths:
        if p and os.path.isfile(p):
            return p
    return None

def recalc(filepath, timeout=30):
    """Recalculate formulas and check for errors."""
    soffice = find_libreoffice()
    if not soffice:
        return {"status": "error", "message": "LibreOffice not found. Install it to recalculate formulas."}

    filepath = os.path.abspath(filepath)
    if not os.path.isfile(filepath):
        return {"status": "error", "message": f"File not found: {filepath}"}

    # Use LibreOffice to recalculate
    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            result = subprocess.run(
                [soffice, "--headless", "--calc", "--convert-to", "xlsx", filepath, "--outdir", tmpdir],
                capture_output=True, text=True, timeout=timeout
            )
            # Copy recalculated file back
            out_file = os.path.join(tmpdir, os.path.basename(filepath))
            if os.path.isfile(out_file):
                shutil.copy2(out_file, filepath)
        except subprocess.TimeoutExpired:
            return {"status": "error", "message": f"LibreOffice timed out after {timeout}s"}
        except Exception as e:
            return {"status": "error", "message": str(e)}

    # Check for formula errors using openpyxl
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        errors = {}
        total_formulas = 0
        error_types = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value in error_types:
                        err = cell.value
                        if err not in errors:
                            errors[err] = {"count": 0, "locations": []}
                        errors[err]["count"] += 1
                        errors[err]["locations"].append(f"{sheet_name}!{cell.coordinate}")

        wb_formulas = load_workbook(filepath)
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        total_formulas += 1

        total_errors = sum(e["count"] for e in errors.values())
        return {
            "status": "errors_found" if total_errors > 0 else "success",
            "total_errors": total_errors,
            "total_formulas": total_formulas,
            "error_summary": errors if errors else {}
        }
    except ImportError:
        return {"status": "warning", "message": "openpyxl not installed. Cannot verify formulas."}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout]")
        sys.exit(1)
    
    filepath = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    result = recalc(filepath, timeout)
    print(json.dumps(result, indent=2))
