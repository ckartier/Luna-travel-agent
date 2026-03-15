#!/usr/bin/env python3
"""
Generate a Luna-branded XLSX quote from JSON input.
Usage: echo '{"quote": {...}}' | python3 generate_quote_xlsx.py /tmp/output.xlsx
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Luna brand colors
LUNA_CHARCOAL = "2E2E2E"
LUNA_PRIMARY = "BCDEEA"
LUNA_BG = "F8FAFC"
LUNA_BORDER = "E5E7EB"
LUNA_MUTED = "6B7280"
LUNA_WHITE = "FFFFFF"

# Styles
HEADER_FONT = Font(name="Arial", size=14, bold=True, color=LUNA_WHITE)
TITLE_FONT = Font(name="Arial", size=11, bold=True, color=LUNA_CHARCOAL)
BODY_FONT = Font(name="Arial", size=10, color=LUNA_CHARCOAL)
MUTED_FONT = Font(name="Arial", size=9, color=LUNA_MUTED)
MONEY_FONT = Font(name="Arial", size=10, bold=True, color=LUNA_CHARCOAL)
TOTAL_FONT = Font(name="Arial", size=12, bold=True, color=LUNA_WHITE)

HEADER_FILL = PatternFill("solid", fgColor=LUNA_CHARCOAL)
PRIMARY_FILL = PatternFill("solid", fgColor=LUNA_PRIMARY)
BG_FILL = PatternFill("solid", fgColor=LUNA_BG)
WHITE_FILL = PatternFill("solid", fgColor=LUNA_WHITE)
TOTAL_FILL = PatternFill("solid", fgColor=LUNA_CHARCOAL)

THIN_BORDER = Border(
    bottom=Side(style="thin", color=LUNA_BORDER),
)
ALL_BORDER = Border(
    left=Side(style="thin", color=LUNA_BORDER),
    right=Side(style="thin", color=LUNA_BORDER),
    top=Side(style="thin", color=LUNA_BORDER),
    bottom=Side(style="thin", color=LUNA_BORDER),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

EUR_FORMAT = '#,##0.00 €'


def generate(data, output_path):
    quote = data.get("quote", {})
    items = quote.get("items", [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Devis {quote.get('quoteNumber', '')}"

    # Column widths
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18

    # ═══ HEADER BLOCK ═══
    row = 2
    ws.merge_cells(f'B{row}:F{row}')
    cell = ws[f'B{row}']
    cell.value = "LUNA CONCIERGERIE"
    cell.font = Font(name="Arial", size=18, bold=True, color=LUNA_CHARCOAL)
    cell.alignment = LEFT

    row = 3
    ws.merge_cells(f'B{row}:F{row}')
    cell = ws[f'B{row}']
    cell.value = "Travel beautifully."
    cell.font = Font(name="Arial", size=10, italic=True, color=LUNA_MUTED)
    cell.alignment = LEFT

    # ═══ QUOTE INFO ═══
    row = 5
    ws[f'B{row}'].value = "Devis N°"
    ws[f'B{row}'].font = MUTED_FONT
    ws[f'C{row}'].value = quote.get("quoteNumber", "")
    ws[f'C{row}'].font = TITLE_FONT

    ws[f'E{row}'].value = "Date :"
    ws[f'E{row}'].font = MUTED_FONT
    ws[f'E{row}'].alignment = RIGHT
    ws[f'F{row}'].value = quote.get("issueDate", "")
    ws[f'F{row}'].font = BODY_FONT

    row = 6
    ws[f'B{row}'].value = "Client"
    ws[f'B{row}'].font = MUTED_FONT
    ws[f'C{row}'].value = quote.get("clientName", "")
    ws[f'C{row}'].font = TITLE_FONT

    ws[f'E{row}'].value = "Valide jusqu'au :"
    ws[f'E{row}'].font = MUTED_FONT
    ws[f'E{row}'].alignment = RIGHT
    ws[f'F{row}'].value = quote.get("validUntil", "")
    ws[f'F{row}'].font = BODY_FONT

    # ═══ TABLE HEADER ═══
    row = 8
    headers = ["#", "Désignation", "Qté", "Coût Net (€)", "Prix Vente (€)", "Total (€)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(name="Arial", size=9, bold=True, color=LUNA_WHITE)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER if col_idx != 2 else LEFT
        cell.border = ALL_BORDER

    # ═══ TABLE BODY ═══
    start_row = row + 1
    for idx, item in enumerate(items, 1):
        r = start_row + idx - 1
        fill = WHITE_FILL if idx % 2 == 1 else BG_FILL

        ws.cell(row=r, column=1, value=idx).font = MUTED_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).border = ALL_BORDER

        ws.cell(row=r, column=2, value=item.get("description", "")).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).border = ALL_BORDER

        qty = item.get("quantity", 1)
        ws.cell(row=r, column=3, value=qty).font = BODY_FONT
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=3).fill = fill
        ws.cell(row=r, column=3).border = ALL_BORDER

        net = item.get("netCost", 0)
        ws.cell(row=r, column=4, value=net).font = MUTED_FONT
        ws.cell(row=r, column=4).number_format = EUR_FORMAT
        ws.cell(row=r, column=4).alignment = RIGHT
        ws.cell(row=r, column=4).fill = fill
        ws.cell(row=r, column=4).border = ALL_BORDER

        price = item.get("unitPrice", 0)
        ws.cell(row=r, column=5, value=price).font = MONEY_FONT
        ws.cell(row=r, column=5).number_format = EUR_FORMAT
        ws.cell(row=r, column=5).alignment = RIGHT
        ws.cell(row=r, column=5).fill = fill
        ws.cell(row=r, column=5).border = ALL_BORDER

        # Total formula: Qty × Unit Price
        col_c = get_column_letter(3)
        col_e = get_column_letter(5)
        ws.cell(row=r, column=6).value = f'={col_c}{r}*{col_e}{r}'
        ws.cell(row=r, column=6).font = MONEY_FONT
        ws.cell(row=r, column=6).number_format = EUR_FORMAT
        ws.cell(row=r, column=6).alignment = RIGHT
        ws.cell(row=r, column=6).fill = fill
        ws.cell(row=r, column=6).border = ALL_BORDER

    end_row = start_row + len(items) - 1

    # ═══ TOTALS ═══
    gap = end_row + 2

    # Subtotal with formula
    ws.merge_cells(f'B{gap}:E{gap}')
    ws[f'B{gap}'].value = "Sous-total HT"
    ws[f'B{gap}'].font = TITLE_FONT
    ws[f'B{gap}'].alignment = RIGHT
    ws[f'F{gap}'].value = f'=SUM(F{start_row}:F{end_row})'
    ws[f'F{gap}'].font = MONEY_FONT
    ws[f'F{gap}'].number_format = EUR_FORMAT
    ws[f'F{gap}'].alignment = RIGHT
    ws[f'F{gap}'].border = THIN_BORDER

    # Tax
    tax_rate = quote.get("taxRate", 20) if quote.get("taxRate") else 20
    tax_row = gap + 1
    ws.merge_cells(f'B{tax_row}:E{tax_row}')
    ws[f'B{tax_row}'].value = f"TVA ({tax_rate}%)"
    ws[f'B{tax_row}'].font = MUTED_FONT
    ws[f'B{tax_row}'].alignment = RIGHT
    ws[f'F{tax_row}'].value = f'=F{gap}*{tax_rate/100}'
    ws[f'F{tax_row}'].font = MUTED_FONT
    ws[f'F{tax_row}'].number_format = EUR_FORMAT
    ws[f'F{tax_row}'].alignment = RIGHT
    ws[f'F{tax_row}'].border = THIN_BORDER

    # Total TTC
    total_row = tax_row + 1
    ws.merge_cells(f'B{total_row}:E{total_row}')
    ws[f'B{total_row}'].value = "TOTAL TTC"
    ws[f'B{total_row}'].font = TOTAL_FONT
    ws[f'B{total_row}'].fill = TOTAL_FILL
    ws[f'B{total_row}'].alignment = Alignment(horizontal="right", vertical="center")
    # Apply fill to merged cells
    for col in range(3, 6):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
    ws[f'F{total_row}'].value = f'=F{gap}+F{tax_row}'
    ws[f'F{total_row}'].font = TOTAL_FONT
    ws[f'F{total_row}'].fill = TOTAL_FILL
    ws[f'F{total_row}'].number_format = EUR_FORMAT
    ws[f'F{total_row}'].alignment = RIGHT

    # ═══ MARGIN ANALYSIS (internal) ═══
    margin_row = total_row + 2
    ws[f'B{margin_row}'].value = "📊 Analyse Interne (non visible client)"
    ws[f'B{margin_row}'].font = Font(name="Arial", size=9, bold=True, color=LUNA_MUTED)

    m1 = margin_row + 1
    ws[f'B{m1}'].value = "Coût net total"
    ws[f'B{m1}'].font = MUTED_FONT
    ws[f'D{m1}'].value = f'=SUMPRODUCT(C{start_row}:C{end_row},D{start_row}:D{end_row})'
    ws[f'D{m1}'].font = MUTED_FONT
    ws[f'D{m1}'].number_format = EUR_FORMAT

    m2 = margin_row + 2
    ws[f'B{m2}'].value = "Marge brute"
    ws[f'B{m2}'].font = Font(name="Arial", size=10, bold=True, color="16A34A")
    ws[f'D{m2}'].value = f'=F{gap}-D{m1}'
    ws[f'D{m2}'].font = Font(name="Arial", size=10, bold=True, color="16A34A")
    ws[f'D{m2}'].number_format = EUR_FORMAT

    m3 = margin_row + 3
    ws[f'B{m3}'].value = "Taux de marge"
    ws[f'B{m3}'].font = MUTED_FONT
    ws[f'D{m3}'].value = f'=IF(F{gap}>0,D{m2}/F{gap},0)'
    ws[f'D{m3}'].font = Font(name="Arial", size=10, bold=True, color="16A34A")
    ws[f'D{m3}'].number_format = '0.0%'

    # ═══ FOOTER ═══
    footer_row = m3 + 3
    ws.merge_cells(f'B{footer_row}:F{footer_row}')
    ws[f'B{footer_row}'].value = "Luna Conciergerie — www.luna-conciergerie.com — Travel beautifully."
    ws[f'B{footer_row}'].font = Font(name="Arial", size=8, italic=True, color=LUNA_MUTED)
    ws[f'B{footer_row}'].alignment = Alignment(horizontal="center")

    # Print settings
    ws.sheet_properties.pageSetUpPr = None
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

    wb.save(output_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: echo '{...}' | python3 generate_quote_xlsx.py output.xlsx", file=sys.stderr)
        sys.exit(1)

    output_path = sys.argv[1]
    data = json.loads(sys.stdin.read())
    generate(data, output_path)
    print(f"Generated: {output_path}", file=sys.stderr)
